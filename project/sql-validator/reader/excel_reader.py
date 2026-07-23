from pathlib import Path
from openpyxl import load_workbook
from common.app_config import get_master_columns,worksheet_to_remove
from common.logger import setup_logger

from services.likely import SimilarColumn

class ExcelReader:
    # clean workbook after removing unwanted column.
    excel_sheet = None
    def __init__(self, file_path: str):
        self.logger = setup_logger()
        self.file_path = Path(file_path)
        #* using config file to load the master columns.
        self.master_columns = get_master_columns()
        # self.user_info()
        
    def user_info(self):
        print("Checkiing for:\n")
        for e_col, e_type in self.master_columns.items():
            print(f"{e_col} '->' {e_type}")

    def read(self):
        self.logger.info(f"Reading Excel file. Path: {self.file_path}\n")

        workbook = load_workbook(self.file_path)
        # check the sheets names - sanity check
        # print(workbook.sheetnames)

        # sheet_name = "Table Name"

        # if unwanted excel sheets are present, we can remove  or filter them and create an updated excel
        sheet_to_removed = worksheet_to_remove()
        if sheet_to_removed:
            for sheet_name in sheet_to_removed:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    workbook.remove(worksheet)

        # save workbook after removing the specified columns
        self.excel_sheet = workbook
        workbook.save("data/data_test.xlsx")

        self.validateExcel()
       
    def validateExcel(self):
        cleaned_workbook = self.excel_sheet
        # store missing columns
        missing_columns = []
        # store mis-match data type
        wrong_types = []
        worksheet_schema = {}

        # loop through the sheets found.
        for sheet in cleaned_workbook.worksheets: # pyright: ignore[reportOptionalMemberAccess]

            sheet_title = sheet.title

            # Initialize schema for this sheet
            worksheet_schema[sheet_title] = {}

            # Read columns A and B
            for c_name, dtype in sheet.iter_rows(
                min_col=1,
                max_col=2,
                values_only=True
            ):

                if c_name:
                    worksheet_schema[sheet_title][str(c_name).strip().lower()] = (
                        str(dtype).strip()
                    )
            # store missing columns
            missing_columns = []
            # store mis-match data type
            wrong_types = []

            # get current sheet and store it in sheet_schema
            sheet_schema = worksheet_schema[sheet_title]
            # print(sheet_schema)

            # loop in using our master colmuns to get the expected col and type
            for e_col, e_type in self.master_columns.items():
                # get the data type stored in sheet.
                # need to use .lower() to avoid conflict in key mismatch

                actual_dtype = sheet_schema.get(e_col.lower())
                # print(e_col, actual_dtype, e_type)
                # print("SheetSchema\n", sheet_schema)

                # 1. Initialize variables to track the best match
                best_match = None
                highest_score = 0.0
                # add missing col if not found 
                if actual_dtype is None:
                    res = e_col
                    # Normalize keys and input to lowercase to fix the case-sensitivity issue
                    clean_input = e_col.lower().strip()
                    clean_keys = [k.lower().strip() for k in sheet_schema.keys()]

                    # get similar columns
                    matches = SimilarColumn().getSimilarColumn(sourceWord=clean_input, possibilities=clean_keys)
                    # print(matches)
                    if not matches:
                        res = f"{e_col} '->' {matches}"
                        missing_columns.append(f"{e_col} '->' {matches}")
                    else:
                        res = f"{e_col} '->'{matches}"
                        missing_columns.append(res)
                    continue
                
                # check and add if data type mis match.
                if actual_dtype != e_type.lower():
                    wrong_types.append(
                        (e_col, e_type, actual_dtype)
                    )

            print(f"\n=== Worksheet: {sheet_title} ===")

            if missing_columns:
                print("Missing columns:")
                for col in missing_columns:
                    print(f"  - {col}")

            if wrong_types:
                print("Type mismatches:")
                for col, expected, actual in wrong_types:
                    print(
                        f"  - {col}: expected {expected}, found {actual}"
                    )

            if not missing_columns and not wrong_types:
                print("OK. Schema validation passed.")

        return missing_columns, wrong_types

